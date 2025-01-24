import openpyxl as pyxl


def read_xlsx_file(file:str):
    print(file)
    book_obj = pyxl.load_workbook(file)
    excel_sheet = book_obj.active
    result = excel_sheet.cell(1, 2)
    print(result.value)
    return result.value



def create_excel(filename:str) :
    book_obj = pyxl.Workbook()
    excel_sheet = book_obj.active
    excel_sheet['A1'] = 'Name'
    excel_sheet['A2'] = 'student'
    excel_sheet['B1'] = 'age'
    excel_sheet['B2'] = '24'
    book_obj.save(filename)
    print("Excel created successfully")
    return True

def get_average(file)->int :
    print(file)
    book_obj = pyxl.load_workbook(file)
    range1 = book_obj['Sheet']['B2:B5']
    cellsum = 0
    for i, cell in enumerate(range1, 1):
        #print(i)
        cellsum += cell[0].value
    average = cellsum/i
    print(average)
    return average

